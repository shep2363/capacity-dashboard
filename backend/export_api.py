from __future__ import annotations

import json
import math
import os
import tempfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})


def _env_int(name: str, default: int, minimum: int = 1) -> int:
    raw = os.getenv(name, str(default))
    try:
        return max(minimum, int(raw))
    except (TypeError, ValueError):
        return default


def _env_float(name: str, default: float, minimum: float = 0.0) -> float:
    raw = os.getenv(name, str(default))
    try:
        return max(minimum, float(raw))
    except (TypeError, ValueError):
        return default


DEFAULT_MAX_UPLOAD_BYTES = 30 * 1024 * 1024
MAX_UPLOAD_BYTES = _env_int("CAPACITY_MAX_UPLOAD_BYTES", DEFAULT_MAX_UPLOAD_BYTES)
DEFAULT_MAX_PLANNING_OVERRIDES = 100_000
MAX_PLANNING_OVERRIDES = _env_int("CAPACITY_MAX_PLANNING_OVERRIDES", DEFAULT_MAX_PLANNING_OVERRIDES)
DEFAULT_MAX_OVERRIDE_HOURS = 1_000_000.0
MAX_OVERRIDE_HOURS = _env_float("CAPACITY_MAX_OVERRIDE_HOURS", DEFAULT_MAX_OVERRIDE_HOURS)
DEFAULT_MAX_RATE_PROJECTS = 10_000
MAX_RATE_PROJECTS = _env_int("CAPACITY_MAX_RATE_PROJECTS", DEFAULT_MAX_RATE_PROJECTS)
DEFAULT_MAX_RATE_PER_HOUR = 1_000_000.0
MAX_RATE_PER_HOUR = _env_float("CAPACITY_MAX_RATE_PER_HOUR", DEFAULT_MAX_RATE_PER_HOUR)
DATA_ROOT = Path(os.getenv("CAPACITY_SHARED_DATA_DIR", Path(__file__).resolve().parent / "shared_store")).resolve()
DATASETS = {"main", "sales"}
MANIFEST_DIR = DATA_ROOT / "manifests"


def _json_no_store(payload: Dict[str, Any], status: int = 200) -> Response:
    response = jsonify(payload)
    response.status_code = status
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _default_manifest(dataset: str) -> Dict[str, Any]:
    return {"version": 1, "dataset": dataset, "fileName": "", "uploadedAt": None, "sizeBytes": 0}


def _validate_dataset(dataset: str) -> str:
    normalized = (dataset or "").strip().lower()
    if normalized not in DATASETS:
        raise ValueError("Invalid dataset. Allowed values: main, sales.")
    return normalized


def _active_workbook_path(dataset: str) -> Path:
    return DATA_ROOT / f"active_{dataset}.xlsx"


def _manifest_path(dataset: str) -> Path:
    return MANIFEST_DIR / f"{dataset}.json"


def _planning_state_path(dataset: str) -> Path:
    return DATA_ROOT / f"planning_state_{dataset}.json"


def _revenue_rates_path(dataset: str) -> Path:
    return DATA_ROOT / f"revenue_rates_{dataset}.json"


def _default_planning_state(dataset: str) -> Dict[str, Any]:
    return {
        "dataset": dataset,
        "version": 0,
        "updatedAt": None,
        "source": "system",
        "overrideCount": 0,
        "overrides": {},
    }


def _default_revenue_rates_state(dataset: str) -> Dict[str, Any]:
    return {
        "dataset": dataset,
        "version": 0,
        "updatedAt": None,
        "source": "system",
        "rateCount": 0,
        "rates": {},
    }


def _normalize_source(source: Any) -> str:
    if not isinstance(source, str):
        return "ui"
    trimmed = source.strip()
    if not trimmed:
        return "ui"
    safe = "".join(char if char.isalnum() or char in {"-", "_", "."} else "_" for char in trimmed)
    return safe[:80] or "ui"


def _normalize_base_version(base_version: Any) -> int | None:
    if base_version is None:
        return None
    if isinstance(base_version, bool):
        raise ValueError("baseVersion must be an integer.")
    try:
        parsed = int(base_version)
    except (TypeError, ValueError) as exc:
        raise ValueError("baseVersion must be an integer.") from exc
    if parsed < 0:
        raise ValueError("baseVersion must be greater than or equal to 0.")
    return parsed


def _normalize_overrides(overrides: Any) -> Dict[str, float]:
    if overrides is None:
        return {}
    if not isinstance(overrides, dict):
        raise ValueError("Invalid planning state payload. 'overrides' must be an object.")
    if len(overrides) > MAX_PLANNING_OVERRIDES:
        raise ValueError(f"Planning state exceeds maximum of {MAX_PLANNING_OVERRIDES} override entries.")
    normalized: Dict[str, float] = {}
    for raw_key, raw_value in overrides.items():
        if not isinstance(raw_key, str):
            raise ValueError("Invalid planning override key.")
        key = raw_key.strip()
        if not key:
            raise ValueError("Planning override keys cannot be empty.")
        if len(key) > 512 or "\n" in key or "\r" in key:
            raise ValueError("Invalid planning override key format.")
        try:
            value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid planning override value for key '{key}'.") from exc
        if not math.isfinite(value) or value < 0:
            raise ValueError(f"Planning override value for key '{key}' must be a finite non-negative number.")
        if value > MAX_OVERRIDE_HOURS:
            raise ValueError(f"Planning override value for key '{key}' exceeds maximum of {MAX_OVERRIDE_HOURS}.")
        normalized[key] = value
    return normalized


def _normalize_rate_value(value: Any, project: str, field_name: str) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field_name} for project '{project}'.") from exc
    if not math.isfinite(numeric) or numeric < 0:
        raise ValueError(f"{field_name} for project '{project}' must be a finite non-negative number.")
    if numeric > MAX_RATE_PER_HOUR:
        raise ValueError(f"{field_name} for project '{project}' exceeds maximum of {MAX_RATE_PER_HOUR}.")
    return numeric


def _normalize_revenue_rates(rates: Any) -> Dict[str, Dict[str, float]]:
    if rates is None:
        return {}
    if not isinstance(rates, dict):
        raise ValueError("Invalid revenue rates payload. 'rates' must be an object.")
    if len(rates) > MAX_RATE_PROJECTS:
        raise ValueError(f"Revenue rates exceed maximum of {MAX_RATE_PROJECTS} projects.")
    normalized: Dict[str, Dict[str, float]] = {}
    for raw_project, raw_rate in rates.items():
        if not isinstance(raw_project, str):
            raise ValueError("Invalid revenue rate project key.")
        project = raw_project.strip()
        if not project:
            raise ValueError("Revenue rate project keys cannot be empty.")
        if len(project) > 256 or "\n" in project or "\r" in project:
            raise ValueError("Invalid revenue rate project key format.")
        if not isinstance(raw_rate, dict):
            raise ValueError(f"Invalid revenue rate payload for project '{project}'.")
        revenue_per_hour = _normalize_rate_value(raw_rate.get("revenuePerHour", 0), project, "revenuePerHour")
        gross_profit_per_hour = _normalize_rate_value(
            raw_rate.get("grossProfitPerHour", 0),
            project,
            "grossProfitPerHour",
        )
        normalized[project] = {
            "revenuePerHour": revenue_per_hour,
            "grossProfitPerHour": gross_profit_per_hour,
        }
    return normalized


def _coerce_planning_state(dataset: str, payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return _default_planning_state(dataset)
    version = payload.get("version")
    if isinstance(version, bool) or not isinstance(version, int) or version < 0:
        version = 0
    updated_at = payload.get("updatedAt")
    if not isinstance(updated_at, str):
        updated_at = None
    overrides_raw = payload.get("overrides")
    try:
        overrides = _normalize_overrides(overrides_raw if overrides_raw is not None else {})
    except ValueError:
        overrides = {}
    return {
        "dataset": dataset,
        "version": version,
        "updatedAt": updated_at,
        "source": _normalize_source(payload.get("source")),
        "overrideCount": len(overrides),
        "overrides": overrides,
    }


def _coerce_revenue_rates_state(dataset: str, payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return _default_revenue_rates_state(dataset)
    version = payload.get("version")
    if isinstance(version, bool) or not isinstance(version, int) or version < 0:
        version = 0
    updated_at = payload.get("updatedAt")
    if not isinstance(updated_at, str):
        updated_at = None
    rates_raw = payload.get("rates")
    try:
        rates = _normalize_revenue_rates(rates_raw if rates_raw is not None else {})
    except ValueError:
        rates = {}
    return {
        "dataset": dataset,
        "version": version,
        "updatedAt": updated_at,
        "source": _normalize_source(payload.get("source")),
        "rateCount": len(rates),
        "rates": rates,
    }


def _write_json_atomic(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_path = tempfile.mkstemp(prefix=f"{path.name}.", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=True, separators=(",", ":"))
        os.replace(temp_path, path)
    except Exception:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
        raise


def _read_manifest(dataset: str) -> Dict[str, Any]:
    path = _manifest_path(dataset)
    if not path.exists():
        return _default_manifest(dataset)
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
            if isinstance(payload, dict):
                return payload
    except Exception:
        app.logger.exception("Failed reading manifest.")
    return _default_manifest(dataset)


def _read_planning_state(dataset: str) -> Dict[str, Any]:
    path = _planning_state_path(dataset)
    if not path.exists():
        return _default_planning_state(dataset)
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        app.logger.exception("Failed reading planning state. dataset=%s", dataset)
        return _default_planning_state(dataset)
    return _coerce_planning_state(dataset, payload)


def _read_revenue_rates_state(dataset: str) -> Dict[str, Any]:
    path = _revenue_rates_path(dataset)
    if not path.exists():
        return _default_revenue_rates_state(dataset)
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        app.logger.exception("Failed reading revenue rates state. dataset=%s", dataset)
        return _default_revenue_rates_state(dataset)
    return _coerce_revenue_rates_state(dataset, payload)


def _save_planning_state(dataset: str, overrides: Any, base_version: Any, source: Any) -> Dict[str, Any]:
    normalized = _validate_dataset(dataset)
    normalized_overrides = _normalize_overrides(overrides)
    normalized_base_version = _normalize_base_version(base_version)
    current_state = _read_planning_state(normalized)
    current_version = int(current_state.get("version") or 0)
    if normalized_base_version is not None and normalized_base_version != current_version:
        raise RuntimeError(
            f"Planning state version conflict for dataset '{normalized}'. "
            f"Expected {normalized_base_version}, current is {current_version}."
        )
    next_state = {
        "dataset": normalized,
        "version": current_version + 1,
        "updatedAt": _utc_now_iso(),
        "source": _normalize_source(source),
        "overrideCount": len(normalized_overrides),
        "overrides": normalized_overrides,
    }
    _write_json_atomic(_planning_state_path(normalized), next_state)
    return next_state


def _save_revenue_rates(dataset: str, rates: Any, base_version: Any, source: Any) -> Dict[str, Any]:
    normalized = _validate_dataset(dataset)
    normalized_rates = _normalize_revenue_rates(rates)
    normalized_base_version = _normalize_base_version(base_version)
    current_state = _read_revenue_rates_state(normalized)
    current_version = int(current_state.get("version") or 0)
    if normalized_base_version is not None and normalized_base_version != current_version:
        raise RuntimeError(
            f"Revenue rates version conflict for dataset '{normalized}'. "
            f"Expected {normalized_base_version}, current is {current_version}."
        )
    next_state = {
        "dataset": normalized,
        "version": current_version + 1,
        "updatedAt": _utc_now_iso(),
        "source": _normalize_source(source),
        "rateCount": len(normalized_rates),
        "rates": normalized_rates,
    }
    _write_json_atomic(_revenue_rates_path(normalized), next_state)
    return next_state


def _ensure_store() -> None:
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)
    for dataset in DATASETS:
        path = _manifest_path(dataset)
        if not path.exists():
            _write_json_atomic(path, _default_manifest(dataset))
        planning_path = _planning_state_path(dataset)
        if not planning_path.exists():
            _write_json_atomic(planning_path, _default_planning_state(dataset))
        revenue_path = _revenue_rates_path(dataset)
        if not revenue_path.exists():
            _write_json_atomic(revenue_path, _default_revenue_rates_state(dataset))


def _workbook_state_response(dataset: str) -> Dict[str, Any]:
    normalized = _validate_dataset(dataset)
    workbook_path = _active_workbook_path(normalized)
    manifest = _read_manifest(normalized)
    has_workbook = workbook_path.exists()
    size_bytes = int(manifest.get("sizeBytes") or 0)
    if has_workbook and size_bytes <= 0:
        size_bytes = workbook_path.stat().st_size
    return {
        "dataset": normalized,
        "hasWorkbook": has_workbook,
        "fileName": str(manifest.get("fileName") or (workbook_path.name if has_workbook else "")),
        "uploadedAt": manifest.get("uploadedAt"),
        "sizeBytes": size_bytes if has_workbook else 0,
    }


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 52)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _append_rows(ws, header: List[str], rows: List[List[Any]]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _autosize_columns(ws)


def _build_workbook(payload: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    chart_rows = payload.get("weeklyCapacityChartRows", [])
    weekly_rows = payload.get("weeklyForecastRows", [])
    monthly_rows = payload.get("monthlyForecastRows", [])
    summary_rows = payload.get("summaryRows", [])
    chart_category_keys = payload.get("chartCategoryKeys", [])
    if not isinstance(chart_category_keys, list):
        chart_category_keys = []
    chart_category_keys = [str(key) for key in chart_category_keys]
    chart_category_count = len(chart_category_keys)

    ws_chart = wb.create_sheet("Weekly Capacity Chart")
    _append_rows(
        ws_chart,
        [
            "Week Start",
            "Week End",
            "Week Label",
            "Total Forecast Hours",
            "Capacity",
            "Variance",
            "Status",
            *chart_category_keys,
        ],
        chart_rows,
    )

    if len(chart_rows) > 0 and chart_category_count > 0:
        min_row = 1
        max_row = len(chart_rows) + 1
        cat_start_col = 8
        cat_end_col = cat_start_col + chart_category_count - 1

        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.overlap = 100
        bar.title = "Weekly Capacity Forecast"
        bar.y_axis.title = "Hours"
        bar.x_axis.title = "Week"

        data = Reference(ws_chart, min_col=cat_start_col, max_col=cat_end_col, min_row=min_row, max_row=max_row)
        categories = Reference(ws_chart, min_col=3, min_row=2, max_row=max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)

        line = LineChart()
        line.y_axis.title = "Capacity"
        line_data = Reference(ws_chart, min_col=5, max_col=5, min_row=min_row, max_row=max_row)
        line.add_data(line_data, titles_from_data=True)
        line.set_categories(categories)
        line.y_axis.axId = 200
        line.x_axis = bar.x_axis

        bar += line
        bar.width = 28
        bar.height = 12
        ws_chart.add_chart(bar, "A3")

    ws_weekly = wb.create_sheet("Weekly Forecast")
    _append_rows(
        ws_weekly,
        ["Week Start", "Week End", "Week Label", "Forecast Hours", "Capacity", "Variance", "Status"],
        weekly_rows,
    )

    ws_monthly = wb.create_sheet("Monthly Forecast")
    _append_rows(
        ws_monthly,
        ["Month", "Planned Hours", "Capacity", "Variance", "Status"],
        monthly_rows,
    )

    ws_summary = wb.create_sheet("Summary")
    _append_rows(ws_summary, ["Metric", "Value"], summary_rows)

    return wb


@app.get("/api/workbook-state")
def workbook_state() -> Response:
    dataset = request.args.get("dataset", "main")
    try:
        return _json_no_store(_workbook_state_response(dataset))
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)


@app.get("/api/workbook-file")
def workbook_file() -> Response:
    dataset = request.args.get("dataset", "main")
    try:
        normalized = _validate_dataset(dataset)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)

    workbook_path = _active_workbook_path(normalized)
    if not workbook_path.exists():
        return _json_no_store({"error": f"No workbook has been uploaded for dataset '{normalized}'."}, 404)

    manifest = _read_manifest(normalized)
    download_name = str(manifest.get("fileName") or workbook_path.name)
    response = send_file(
        workbook_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=False,
        download_name=download_name,
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


@app.post("/api/upload-workbook")
def upload_workbook() -> Response:
    dataset = request.args.get("dataset", "main")
    try:
        normalized = _validate_dataset(dataset)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)

    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return _json_no_store({"error": "Missing upload file. Use form field 'file' with a .xlsx workbook."}, 400)

    original_name = Path(upload.filename).name
    if not original_name.lower().endswith(".xlsx"):
        return _json_no_store({"error": "Only .xlsx files are allowed."}, 400)

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, dir=str(DATA_ROOT), prefix="active_", suffix=".tmp") as temp_file:
            upload.save(temp_file)
            temp_path = Path(temp_file.name)

        if temp_path is None or not temp_path.exists():
            return _json_no_store({"error": "Saving the workbook file failed."}, 500)

        file_size = temp_path.stat().st_size
        if file_size <= 0:
            temp_path.unlink(missing_ok=True)
            return _json_no_store({"error": "Uploaded workbook is empty."}, 400)

        if file_size > MAX_UPLOAD_BYTES:
            temp_path.unlink(missing_ok=True)
            return _json_no_store({"error": f"Workbook exceeds maximum size of {MAX_UPLOAD_BYTES} bytes."}, 413)

        with temp_path.open("rb") as handle:
            if handle.read(2) != b"PK":
                temp_path.unlink(missing_ok=True)
                return _json_no_store({"error": "Invalid workbook payload. Expected .xlsx content."}, 400)

        os.replace(temp_path, _active_workbook_path(normalized))
        uploaded_at = _utc_now_iso()

        manifest = {
            "version": 1,
            "dataset": normalized,
            "fileName": original_name,
            "uploadedAt": uploaded_at,
            "sizeBytes": file_size,
        }
        _write_json_atomic(_manifest_path(normalized), manifest)

        return _json_no_store(
            {
                "dataset": normalized,
                "hasWorkbook": True,
                "fileName": original_name,
                "uploadedAt": uploaded_at,
                "sizeBytes": file_size,
            }
        )
    except Exception:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        app.logger.exception("Failed uploading active workbook.")
        return _json_no_store({"error": "Saving the workbook file failed."}, 500)


@app.get("/api/planning-state")
def get_planning_state() -> Response:
    dataset = request.args.get("dataset", "main")
    try:
        normalized = _validate_dataset(dataset)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)
    return _json_no_store(_read_planning_state(normalized))


@app.post("/api/planning-state")
def save_planning_state() -> Response:
    dataset = request.args.get("dataset", "main")
    payload = request.get_json(silent=True)
    if payload is None:
        payload = {}
    if not isinstance(payload, dict):
        return _json_no_store({"error": "Invalid JSON payload."}, 400)

    overrides = payload.get("overrides", {})
    base_version = payload.get("baseVersion")
    source = payload.get("source", "ui")

    try:
        saved = _save_planning_state(dataset, overrides, base_version, source)
        return _json_no_store(saved)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)
    except RuntimeError as exc:
        return _json_no_store({"error": str(exc)}, 409)
    except Exception:
        app.logger.exception("Failed saving planning state. dataset=%s", dataset)
        return _json_no_store({"error": "Failed saving planning state."}, 500)


@app.get("/api/revenue-rates")
def get_revenue_rates() -> Response:
    dataset = request.args.get("dataset", "main")
    try:
        normalized = _validate_dataset(dataset)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)
    return _json_no_store(_read_revenue_rates_state(normalized))


@app.post("/api/revenue-rates")
def save_revenue_rates() -> Response:
    dataset = request.args.get("dataset", "main")
    payload = request.get_json(silent=True)
    if payload is None:
        payload = {}
    if not isinstance(payload, dict):
        return _json_no_store({"error": "Invalid JSON payload."}, 400)

    rates = payload.get("rates", {})
    base_version = payload.get("baseVersion")
    source = payload.get("source", "ui")

    try:
        saved = _save_revenue_rates(dataset, rates, base_version, source)
        return _json_no_store(saved)
    except ValueError as exc:
        return _json_no_store({"error": str(exc)}, 400)
    except RuntimeError as exc:
        return _json_no_store({"error": str(exc)}, 409)
    except Exception:
        app.logger.exception("Failed saving revenue rates. dataset=%s", dataset)
        return _json_no_store({"error": "Failed saving revenue rates."}, 500)


@app.get("/api/shared-health")
def shared_health() -> Response:
    main_state = _workbook_state_response("main")
    sales_state = _workbook_state_response("sales")
    main_planning = _read_planning_state("main")
    sales_planning = _read_planning_state("sales")
    main_revenue = _read_revenue_rates_state("main")
    sales_revenue = _read_revenue_rates_state("sales")
    return jsonify(
        {
            "status": "ok",
            "dataRoot": str(DATA_ROOT),
            "datasets": {
                "main": main_state,
                "sales": sales_state,
            },
            "planningState": {
                "main": {
                    "version": main_planning.get("version", 0),
                    "updatedAt": main_planning.get("updatedAt"),
                    "overrideCount": main_planning.get("overrideCount", 0),
                },
                "sales": {
                    "version": sales_planning.get("version", 0),
                    "updatedAt": sales_planning.get("updatedAt"),
                    "overrideCount": sales_planning.get("overrideCount", 0),
                },
            },
            "revenueRates": {
                "main": {
                    "version": main_revenue.get("version", 0),
                    "updatedAt": main_revenue.get("updatedAt"),
                    "rateCount": main_revenue.get("rateCount", 0),
                },
                "sales": {
                    "version": sales_revenue.get("version", 0),
                    "updatedAt": sales_revenue.get("updatedAt"),
                    "rateCount": sales_revenue.get("rateCount", 0),
                },
            },
        }
    )


@app.post("/api/export-report")
def export_report() -> Response:
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "Invalid JSON payload"}), 400

    file_name = str(payload.get("fileName") or "capacity-report.xlsx")
    if not file_name.endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"

    workbook = _build_workbook(payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


_ensure_store()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
