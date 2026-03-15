from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

import sys
from pathlib import Path

from flask import Flask, Response, jsonify, request
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Ensure sibling helper modules inside /api are importable in serverless runtime.
sys.path.append(str(Path(__file__).resolve().parent))
from _auth import READ_ONLY_ROLES, require_auth

app = Flask(__name__)


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 52)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _append_rows(ws, header: List[str], rows: List[List[Any]]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _autosize_columns(ws)


def _build_workbook(payload: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    chart_rows = payload.get("weeklyCapacityChartRows", [])
    weekly_rows = payload.get("weeklyForecastRows", [])
    monthly_rows = payload.get("monthlyForecastRows", [])
    summary_rows = payload.get("summaryRows", [])
    chart_category_keys = payload.get("chartCategoryKeys", [])
    if not isinstance(chart_category_keys, list):
        chart_category_keys = []
    chart_category_keys = [str(key) for key in chart_category_keys]
    chart_category_count = len(chart_category_keys)

    ws_chart = wb.create_sheet("Weekly Capacity Chart")
    _append_rows(
        ws_chart,
        [
            "Week Start",
            "Week End",
            "Week Label",
            "Total Forecast Hours",
            "Capacity",
            "Variance",
            "Status",
            *chart_category_keys,
        ],
        chart_rows,
    )

    if len(chart_rows) > 0 and chart_category_count > 0:
        min_row = 1
        max_row = len(chart_rows) + 1
        cat_start_col = 8
        cat_end_col = cat_start_col + chart_category_count - 1

        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.overlap = 100
        bar.title = "Weekly Capacity Forecast"
        bar.y_axis.title = "Hours"
        bar.x_axis.title = "Week"

        data = Reference(ws_chart, min_col=cat_start_col, max_col=cat_end_col, min_row=min_row, max_row=max_row)
        categories = Reference(ws_chart, min_col=3, min_row=2, max_row=max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)

        line = LineChart()
        line.y_axis.title = "Capacity"
        line_data = Reference(ws_chart, min_col=5, max_col=5, min_row=min_row, max_row=max_row)
        line.add_data(line_data, titles_from_data=True)
        line.set_categories(categories)
        line.y_axis.axId = 200
        line.x_axis = bar.x_axis

        bar += line
        bar.width = 28
        bar.height = 12
        ws_chart.add_chart(bar, "A3")

    ws_weekly = wb.create_sheet("Weekly Forecast")
    _append_rows(
        ws_weekly,
        ["Week Start", "Week End", "Week Label", "Forecast Hours", "Capacity", "Variance", "Status"],
        weekly_rows,
    )

    ws_monthly = wb.create_sheet("Monthly Forecast")
    _append_rows(
        ws_monthly,
        ["Month", "Planned Hours", "Capacity", "Variance", "Status"],
        monthly_rows,
    )

    ws_summary = wb.create_sheet("Summary")
    _append_rows(ws_summary, ["Metric", "Value"], summary_rows)

    return wb


@app.post("/")
@app.post("/api/export-report")
@require_auth(READ_ONLY_ROLES)
def export_report() -> Response:
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "Invalid JSON payload"}), 400

    file_name = str(payload.get("fileName") or "capacity-report.xlsx")
    if not file_name.endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"

    workbook = _build_workbook(payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )
